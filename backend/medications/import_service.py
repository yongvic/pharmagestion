import csv
import io
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils import timezone

from .models import Category, Medication

COLUMN_ALIASES = {
    'name': {
        'nom_commercial', 'nom', 'name', 'medicament', 'médicament', 'produit',
    },
    'generic_name': {'nom_generique', 'nom_générique', 'generic_name'},
    'category': {'categorie', 'catégorie', 'category'},
    'dosage': {'dosage'},
    'form': {'forme', 'forme_pharmaceutique', 'form'},
    'manufacturer': {'fabricant', 'manufacturer'},
    'purchase_price': {'prix_achat', 'purchase_price', 'pa'},
    'selling_price': {'prix_vente', 'selling_price', 'pv'},
    'stock_quantity': {'stock', 'quantite', 'quantité', 'stock_quantity'},
    'min_stock': {'stock_minimum', 'stock_min', 'min_stock'},
    'lot_number': {'numero_lot', 'numéro_lot', 'lot', 'lot_number'},
    'expiry_date': {'date_expiration', 'expiration', 'expiry_date', 'date_exp'},
    'barcode': {'code_barres', 'code-barres', 'barcode', 'ean'},
    'location': {'emplacement', 'location'},
    'description': {'description'},
}

REQUIRED_FIELDS = ('name', 'category', 'purchase_price', 'selling_price', 'expiry_date')

TEMPLATE_HEADERS = [
    'nom_commercial',
    'nom_generique',
    'categorie',
    'dosage',
    'forme',
    'fabricant',
    'prix_achat',
    'prix_vente',
    'stock',
    'stock_minimum',
    'numero_lot',
    'date_expiration',
    'code_barres',
    'emplacement',
    'description',
]

TEMPLATE_SAMPLE = [
    'Paracétamol 500mg',
    'Paracétamol',
    'Analgésiques',
    '500mg',
    'Comprimé',
    'Pharma Togo',
    '150',
    '250',
    '100',
    '10',
    'LOT-2026-01',
    '2027-12-31',
    '3760123456789',
    'Rayon A1',
    'Antalgique et antipyrétique',
]

TEMPLATE_SAMPLE_2 = [
    'Amoxicilline 500mg',
    'Amoxicilline',
    'Antibiotiques',
    '500mg',
    'Gélule',
    'Pharma Togo',
    '800',
    '1200',
    '50',
    '10',
    'LOT-2026-02',
    '2028-06-30',
    '3760987654321',
    'Rayon B2',
    'Antibiotique à large spectre',
]

DEFAULT_CATEGORY_EXAMPLES = [
    ('Analgésiques', 'Médicaments contre la douleur et la fièvre'),
    ('Antibiotiques', 'Traitement des infections bactériennes'),
    ('Vitamines', 'Compléments et vitamines'),
    ('Antipaludéens', 'Traitement et prévention du paludisme'),
    ('Dermatologie', 'Soins de la peau'),
]

# Format inventaire hôpital (MEG / CHP / NTG) — ex. "Prix MEG CHP NTG 2026.xlsx"
MEG_SKIP_LABELS = {
    'TOTAL', 'TOTAL GLOBAL', 'COMPTABLE', 'STOCK INITIAL',
    'MAG', 'PHAR', 'DETAIL', 'DESIGATTION', 'DESIGNATION',
}
MEG_FOOTER_EXACT = {
    'TOTAL', 'TOTAL GLOBAL', 'COMPTABLE', 'DIRECTEUR', 'MAGASINIER',
    'PHARMACIEN', 'CHEF PHARMACIE', 'ORDONNATEUR',
}
MEG_SECTION_FORMS = {
    'COMPRIM': 'Comprimé',
    'INJECT': 'Injectable',
    'SIROP': 'Sirop',
    'CONSOMMABLE': 'Consommable',
}
DOSAGE_RE = re.compile(
    r'(\d+[\s,]*(?:mg|ml|g|µg|mcg|%)(?:\s*/\s*[\d,\.]+\s*(?:mg|ml|g)?)?(?:\s*[\w/]+)?)',
    re.IGNORECASE,
)


def normalize_header(value):
    text = str(value or '').strip().lower()
    for old, new in (('é', 'e'), ('è', 'e'), ('ê', 'e'), ('à', 'a'), ('ù', 'u'), ('ô', 'o')):
        text = text.replace(old, new)
    return text.replace(' ', '_').replace('-', '_')


def build_header_map(headers):
    normalized = [normalize_header(h) for h in headers]
    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases or header == field:
                mapping[field] = index
                break
    return mapping


def cell_value(row, mapping, field, default=''):
    index = mapping.get(field)
    if index is None or index >= len(row):
        return default
    value = row[index]
    if value is None:
        return default
    return str(value).strip() if not isinstance(value, (int, float, Decimal)) else value


def parse_decimal(value, field_label):
    if value in ('', None):
        raise ValueError(f'{field_label} est obligatoire.')
    text = str(value).strip().replace(' ', '').replace(',', '.')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f'{field_label} invalide : {value}') from exc


def parse_int(value, default=0):
    if value in ('', None):
        return default
    text = str(value).strip().replace(' ', '')
    try:
        return int(float(text))
    except (ValueError, TypeError) as exc:
        raise ValueError(f'Quantité invalide : {value}') from exc


def parse_date(value, required=True):
    if value in ('', None):
        if required:
            raise ValueError("Date d'expiration obligatoire.")
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Date invalide : {value}. Utilisez AAAA-MM-JJ ou JJ/MM/AAAA.")


def parse_decimal_optional(value, field_label):
    if value in ('', None):
        return None
    text = str(value).strip().replace(' ', '').replace(',', '.')
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f'{field_label} invalide : {value}') from exc


def split_name_dosage(raw_name):
    cleaned = ' '.join(str(raw_name or '').split())
    if not cleaned:
        return '', None
    match = DOSAGE_RE.search(cleaned)
    if match:
        dosage = match.group(1).strip()
        name = cleaned[:match.start()].strip(' ,-') or cleaned
        return name, dosage
    return cleaned, None


def infer_form_from_category(category_name):
    upper = str(category_name or '').upper()
    for key, form in MEG_SECTION_FORMS.items():
        if key in upper:
            return form
    return None


def normalize_meg_label(value):
    return ' '.join(str(value or '').split()).strip()


def is_meg_section_row(row):
    label = normalize_meg_label(row[0] if row else None)
    if not label or len(label) < 3:
        return False
    upper = label.upper()
    if upper.startswith('A LA DATE'):
        return False
    if any(skip in upper for skip in MEG_SKIP_LABELS):
        return False
    has_product_data = any(row[i] not in (None, '') for i in (1, 2, 7, 8, 10, 11) if i < len(row))
    return not has_product_data


def is_meg_format_rows(rows):
    for row in rows[:30]:
        cells = [normalize_header(c) for c in row if c is not None]
        if 'desigattion' in cells or 'designation' in cells:
            if any('peremp' in c for c in cells) and any('cdmt' in c for c in cells):
                return True
    return False


def detect_meg_header_row(rows):
    for index, row in enumerate(rows):
        cells = [normalize_header(c) for c in row if c is not None]
        if ('desigattion' in cells or 'designation' in cells) and any('peremp' in c for c in cells):
            return index
    return None


def resolve_unit_price(unit_value, box_value, pack_size, field_label):
    unit = parse_decimal_optional(unit_value, field_label)
    if unit and unit > 0:
        return unit
    box = parse_decimal_optional(box_value, field_label)
    pack = parse_int(pack_size, 0) if pack_size not in (None, '') else 0
    if box and box > 0 and pack > 0:
        return (box / Decimal(pack)).quantize(Decimal('0.01'))
    if box and box > 0:
        return box
    return Decimal('1')


def parse_meg_row(row, current_category, category_cache, stats):
    raw_name = normalize_meg_label(row[0] if row else None)
    if not raw_name:
        raise ValueError('Nom commercial manquant.')

    name, dosage = split_name_dosage(raw_name)
    if not name:
        raise ValueError('Nom commercial manquant.')

    pack_size = row[2] if len(row) > 2 else None
    stock_phar = parse_int(row[5], 0) if len(row) > 5 else 0
    stock_detail = parse_int(row[6], 0) if len(row) > 6 else 0
    stock_quantity = stock_phar if stock_phar > 0 else stock_detail

    purchase_price = resolve_unit_price(
        row[8] if len(row) > 8 else None,
        row[7] if len(row) > 7 else None,
        pack_size,
        'Prix achat',
    )
    selling_price = resolve_unit_price(
        row[11] if len(row) > 11 else None,
        row[10] if len(row) > 10 else None,
        pack_size,
        'Prix vente',
    )

    expiry = parse_date(row[1] if len(row) > 1 else None, required=False)
    if expiry is None:
        expiry = (timezone.now().date() + timedelta(days=365))

    category_name = current_category or 'Non classé'
    form = infer_form_from_category(category_name)

    return {
        'name': name,
        'generic_name': None,
        'category': resolve_category(category_name, category_cache, stats),
        'dosage': dosage,
        'form': form,
        'manufacturer': None,
        'purchase_price': purchase_price,
        'selling_price': selling_price,
        'stock_quantity': stock_quantity,
        'min_stock': 5,
        'lot_number': None,
        'expiry_date': expiry,
        'barcode': None,
        'location': 'PHAR',
        'description': f'Import inventaire MEG/CHP — {raw_name}',
    }


def parse_meg_sheet_rows(rows):
    header_index = detect_meg_header_row(rows)
    if header_index is None:
        raise ValueError('Feuille inventaire MEG/CHP non reconnue.')

    parsed_rows = []
    current_category = 'Non classé'

    for line_no, row in enumerate(rows[header_index + 2:], start=header_index + 3):
        row = list(row or [])
        label = normalize_meg_label(row[0] if row else None)
        if not label:
            continue

        upper = label.upper()
        if upper in MEG_FOOTER_EXACT or upper.startswith('A LA DATE'):
            continue

        if is_meg_section_row(row):
            current_category = label
            continue

        parsed_rows.append((line_no, row, current_category))

    return parsed_rows


def load_xlsx_workbook(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Support Excel indisponible (openpyxl manquant).') from exc
    return load_workbook(file_obj, read_only=True, data_only=True)


def parse_meg_xlsx_rows(file_obj):
    workbook = load_xlsx_workbook(file_obj)
    sheet = workbook[workbook.sheetnames[-1]]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    sheet_name = sheet.title
    workbook.close()
    if not rows:
        raise ValueError('Le fichier Excel est vide.')
    if not is_meg_format_rows(rows):
        raise ValueError('Format inventaire MEG/CHP non détecté.')
    return sheet_name, parse_meg_sheet_rows(rows)


def find_existing_medication(payload):
    if payload.get('barcode'):
        existing = Medication.objects.filter(barcode=payload['barcode']).first()
        if existing:
            return existing

    queryset = Medication.objects.filter(name__iexact=payload['name'])
    if payload.get('dosage'):
        match = queryset.filter(dosage__iexact=payload['dosage']).first()
        if match:
            return match
    return queryset.first()


def upsert_medication(payload, update_existing, results):
    category = payload.pop('category')
    with transaction.atomic():
        existing = find_existing_medication(payload)
        if existing:
            if not update_existing:
                results['skipped'] += 1
                return
            for key, value in payload.items():
                setattr(existing, key, value)
            existing.category = category
            existing.save()
            results['updated'] += 1
        else:
            Medication.objects.create(category=category, **payload)
            results['created'] += 1


def import_meg_format(file_obj, filename, update_existing=True):
    sheet_name, parsed_rows = parse_meg_xlsx_rows(file_obj)
    results = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'categories_created': 0,
        'errors': [],
        'total_rows': 0,
        'format': 'meg_chp_inventaire',
        'sheet': sheet_name,
    }
    category_cache = {}

    for line_no, row, current_category in parsed_rows:
        results['total_rows'] += 1
        try:
            payload = parse_meg_row(row, current_category, category_cache, results)
            upsert_medication(payload, update_existing, results)
        except Exception as exc:
            results['errors'].append({'line': line_no, 'message': str(exc)})

    return results


def resolve_category(name, cache, stats):
    cleaned = str(name or '').strip()
    if not cleaned:
        raise ValueError('Catégorie obligatoire pour chaque produit.')

    cache_key = cleaned.lower()
    if cache_key in cache:
        return cache[cache_key]

    existing = Category.objects.filter(name__iexact=cleaned).first()
    if existing:
        cache[cache_key] = existing
        return existing

    category = Category.objects.create(name=cleaned)
    cache[cache_key] = category
    stats['categories_created'] += 1
    return category


def peek_xlsx_rows(file_obj):
    workbook = load_xlsx_workbook(file_obj)
    sheet = workbook[workbook.sheetnames[-1]]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    sheet_name = sheet.title
    workbook.close()
    file_obj.seek(0)
    return sheet_name, rows


def parse_csv_rows(file_obj):
    content = file_obj.read()
    if isinstance(content, bytes):
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError('Encodage du fichier CSV non reconnu.')
    else:
        text = content

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError('Le fichier est vide.')
    header_map = build_header_map(rows[0])
    return header_map, rows[1:]


def parse_xlsx_rows(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Support Excel indisponible (openpyxl manquant).') from exc

    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    if not rows:
        raise ValueError('Le fichier Excel est vide.')
    header_map = build_header_map(rows[0])
    return header_map, rows[1:]


def parse_uploaded_file(file_obj, filename):
    ext = filename.rsplit('.', 1)[-1].lower() if filename and '.' in filename else ''
    if ext == 'csv':
        return parse_csv_rows(file_obj)
    if ext == 'xlsx':
        return parse_xlsx_rows(file_obj)
    raise ValueError('Format non supporté. Utilisez .csv ou .xlsx.')


def row_to_payload(row, mapping, category_cache, stats):
    name = str(cell_value(row, mapping, 'name')).strip()
    if not name:
        raise ValueError('Nom commercial manquant.')

    category_name = str(cell_value(row, mapping, 'category')).strip()
    barcode = str(cell_value(row, mapping, 'barcode')).strip() or None

    return {
        'name': name,
        'generic_name': str(cell_value(row, mapping, 'generic_name')).strip() or None,
        'category': resolve_category(category_name, category_cache, stats),
        'dosage': str(cell_value(row, mapping, 'dosage')).strip() or None,
        'form': str(cell_value(row, mapping, 'form')).strip() or None,
        'manufacturer': str(cell_value(row, mapping, 'manufacturer')).strip() or None,
        'purchase_price': parse_decimal(cell_value(row, mapping, 'purchase_price'), 'Prix achat'),
        'selling_price': parse_decimal(cell_value(row, mapping, 'selling_price'), 'Prix vente'),
        'stock_quantity': parse_int(cell_value(row, mapping, 'stock_quantity'), 0),
        'min_stock': parse_int(cell_value(row, mapping, 'min_stock'), 5),
        'lot_number': str(cell_value(row, mapping, 'lot_number')).strip() or None,
        'expiry_date': parse_date(cell_value(row, mapping, 'expiry_date')),
        'barcode': barcode,
        'location': str(cell_value(row, mapping, 'location')).strip() or None,
        'description': str(cell_value(row, mapping, 'description')).strip() or None,
    }


def import_standard_format(file_obj, filename, update_existing=True):
    header_map, data_rows = parse_uploaded_file(file_obj, filename)

    missing = [field for field in REQUIRED_FIELDS if field not in header_map]
    if missing:
        labels = {
            'name': 'nom_commercial',
            'category': 'categorie',
            'purchase_price': 'prix_achat',
            'selling_price': 'prix_vente',
            'expiry_date': 'date_expiration',
        }
        raise ValueError(
            'Colonnes obligatoires manquantes : '
            + ', '.join(labels.get(f, f) for f in missing)
        )

    results = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'categories_created': 0,
        'errors': [],
        'total_rows': 0,
        'format': 'standard',
    }
    category_cache = {}

    for line_no, row in enumerate(data_rows, start=2):
        if not row or all(str(cell or '').strip() == '' for cell in row):
            continue

        results['total_rows'] += 1
        try:
            payload = row_to_payload(row, header_map, category_cache, results)
            upsert_medication(payload, update_existing, results)
        except Exception as exc:
            results['errors'].append({'line': line_no, 'message': str(exc)})

    return results


def import_medications_from_file(file_obj, filename, update_existing=True):
    ext = filename.rsplit('.', 1)[-1].lower() if filename and '.' in filename else ''
    if ext == 'xlsx':
        _, rows = peek_xlsx_rows(file_obj)
        if is_meg_format_rows(rows):
            return import_meg_format(file_obj, filename, update_existing=update_existing)
    return import_standard_format(file_obj, filename, update_existing=update_existing)


def build_template_csv():
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow(TEMPLATE_SAMPLE)
    writer.writerow(TEMPLATE_SAMPLE_2)
    return buffer.getvalue().encode('utf-8-sig')


def build_template_xlsx(categories=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Medicaments'
    sheet.append(TEMPLATE_HEADERS)
    sheet.append(TEMPLATE_SAMPLE)
    sheet.append(TEMPLATE_SAMPLE_2)

    header_fill = PatternFill('solid', fgColor='EEF2FF')
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 18

    cat_sheet = workbook.create_sheet('Categories')
    cat_sheet.append(['categorie', 'description'])
    for cell in cat_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    rows = [(c.name, c.description or '') for c in categories] if categories else DEFAULT_CATEGORY_EXAMPLES
    for name, description in rows:
        cat_sheet.append([name, description])
    cat_sheet.column_dimensions['A'].width = 22
    cat_sheet.column_dimensions['B'].width = 40

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
